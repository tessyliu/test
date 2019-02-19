#-*-coding:utf-8-*-
#@Time    :2019/1/18 13:24
#@Author  :liu_zhenzhen
#@File    :do_excel.py
from openpyxl import load_workbook
class Cases:
    def __init__(self):
        self.case_id=None   #赋空值
        self.module=None
        self.title=None
        self.method=None
        self.url=None
        self.data=None
        self.expectedresult=None

class DoExcel:
    def __init__(self,file_name):
        self.file_name=file_name
        try:
            self.wb = load_workbook(self.file_name)
        except FileNotFoundError as e:  #文件未找到的异常处理
            print("文件未找到")

    def read_excel(self,sheet_name):
        self.sheet_name=sheet_name   #还可以这样？？？不明白？？？？？？
        sheet=self.wb[self.sheet_name]
        cases=[]
        for i in range(2,sheet.max_row+1):
            row_case=Cases()    #每一行数据都存在这个对象中
            row_case.case_id=sheet.cell(row=i,column=1).value
            row_case.module=sheet.cell(row=i,column=2).value
            row_case.title=sheet.cell(row=i,column=3).value
            row_case.method=sheet.cell(row=i,column=4).value
            row_case.url=sheet.cell(row=i,column=5).value
            row_case.data=sheet.cell(row=i,column=6).value
            row_case.expectedresult=sheet.cell(row=i,column=7).value
            if type(row_case.expectedresult) == int:#将期望结果转换成str类型
                row_case.expectedresult = str(row_case.expectedresult)
            cases.append(row_case) #将每一个对象都追加到列表中
        return cases

    def write_back(self,row,actualresult,result):
        sheet =self.wb[self.sheet_name]
        sheet.cell(row,8).value=actualresult
        sheet.cell(row,9).value=result
        self.wb.save(self.file_name)

if __name__=='__main__':
    from common.request_method import RequestMethod
    from common import contants
    import json
    # 注意路径：此处使用的常量文件，另一种使用上级.查找'../datas/cases.xlsx'
    do_excel=DoExcel(contants.cases_dir)
    cases=do_excel.read_excel('register')
    request_1=RequestMethod()
    for case in cases:
        #case.data=eval(case.data) 这一步尽量放在ruqest_method中进行判断
        resp=request_1.request_method(case.method,case.url,case.data)
        if resp.text==case.expectedresult:
            do_excel.write_back(case.case_id+1,resp.text,"Pass")
        else:
            do_excel.write_back(case.case_id+1,resp.text,"Failed")



    # print("case列表：",cases.read_excel())
    # print("用例条数：",len(cases.read_excel()))